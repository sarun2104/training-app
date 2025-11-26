"""
Script to export all courses and their links from FalkorDB to an Excel file
Creates an Excel with two columns: Course and Links
"""
import sys
import os
import logging
from datetime import datetime

# Add parent directory to path to import backend modules
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from backend.database.falkordb import FalkorDB
from backend.config import settings

# Try to import openpyxl, install if not available
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("openpyxl not found. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class CourseLinksExporter:
    """Export courses and their links to Excel"""

    def __init__(self, falkor_db: FalkorDB):
        self.db = falkor_db

    def get_all_courses_with_links(self):
        """Fetch all courses and their associated links from FalkorDB"""
        logger.info("Fetching courses and links from FalkorDB...")

        # First get all courses
        courses_query = """
        MATCH (c:Course)
        RETURN c.course_id as course_id, c.course_name as course_name
        ORDER BY c.course_name
        """

        courses_result = self.db.execute_query(courses_query)

        courses_data = []
        for course_row in courses_result:
            course_id = course_row[0]
            course_name = course_row[1]

            # Get links for this course
            links_query = """
            MATCH (c:Course {course_id: $course_id})-[:has_links]->(l:Links)
            RETURN l.link_label as label, l.link as url
            ORDER BY l.link_label
            """
            links_params = {"course_id": course_id}
            links_result = self.db.execute_query(links_query, links_params)

            # Build links list
            links = []
            for link_row in links_result:
                if link_row[0] and link_row[1]:  # Both label and url exist
                    links.append({
                        'label': link_row[0],
                        'url': link_row[1]
                    })

            courses_data.append({
                'course_name': course_name,
                'links': links
            })

        logger.info(f"Found {len(courses_data)} courses")
        return courses_data

    def create_excel(self, courses_data, output_file):
        """Create Excel file with courses and links"""
        logger.info("Creating Excel file...")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Courses and Links"

        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 100

        # Create header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        ws['A1'] = "Course"
        ws['A1'].fill = header_fill
        ws['A1'].font = header_font
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

        ws['B1'] = "Links"
        ws['B1'].fill = header_fill
        ws['B1'].font = header_font
        ws['B1'].alignment = Alignment(horizontal='left', vertical='center')

        # Add data rows
        row_num = 2
        for course in courses_data:
            course_name = course['course_name']
            links = course['links']

            # Add course name
            ws[f'A{row_num}'] = course_name
            ws[f'A{row_num}'].font = Font(bold=True, size=11)
            ws[f'A{row_num}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            # Add links
            if links:
                links_text = []
                for link in links:
                    label = link['label']
                    url = link['url']
                    links_text.append(f"{label}\n{url}")

                ws[f'B{row_num}'] = "\n\n".join(links_text)
                ws[f'B{row_num}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                # Adjust row height based on number of links
                ws.row_dimensions[row_num].height = max(30, len(links) * 30)
            else:
                ws[f'B{row_num}'] = "No links available"
                ws[f'B{row_num}'].font = Font(italic=True, color="999999")
                ws[f'B{row_num}'].alignment = Alignment(horizontal='left', vertical='top')

            row_num += 1

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Save workbook
        wb.save(output_file)
        logger.info(f"Excel file saved: {output_file}")

    def export_to_excel(self, output_file=None):
        """Main method to export courses and links to Excel"""
        if output_file is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = f"courses_and_links_{timestamp}.xlsx"

        # Get data
        courses_data = self.get_all_courses_with_links()

        # Create Excel
        self.create_excel(courses_data, output_file)

        # Print summary
        total_courses = len(courses_data)
        courses_with_links = sum(1 for course in courses_data if course['links'])
        total_links = sum(len(course['links']) for course in courses_data)

        logger.info("\n" + "="*60)
        logger.info("Export Summary:")
        logger.info("="*60)
        logger.info(f"Total courses: {total_courses}")
        logger.info(f"Courses with links: {courses_with_links}")
        logger.info(f"Courses without links: {total_courses - courses_with_links}")
        logger.info(f"Total links: {total_links}")
        logger.info(f"Average links per course: {total_links / total_courses:.1f}")
        logger.info("="*60)

        return output_file


def main():
    """Main entry point"""
    try:
        # Initialize FalkorDB connection
        logger.info("Connecting to FalkorDB...")
        falkor_db = FalkorDB()
        falkor_db.connect()

        # Create exporter and run
        exporter = CourseLinksExporter(falkor_db)
        output_file = exporter.export_to_excel()

        # Close connection
        falkor_db.close()

        logger.info(f"\n✓ Export completed successfully!")
        logger.info(f"File location: {os.path.abspath(output_file)}")

    except Exception as e:
        logger.error(f"Export failed: {e}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    main()
